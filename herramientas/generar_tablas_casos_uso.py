"""
Genera la documentacion de los casos de uso del sistema NSG Cotizador
en formato XLSX, una hoja por cada caso de uso segun la plantilla RF-01.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---- Estilos ----
TITULO_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITULO_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
CAMPOS_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E78")
CAMPOS_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_TABLA_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_TABLA_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
BORDER_THIN = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def estilizar_titulo(ws, texto, fila=1, columnas=2):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=columnas)
    celda = ws.cell(row=fila, column=1, value=texto)
    celda.font = TITULO_FONT
    celda.fill = TITULO_FILL
    celda.alignment = CENTER
    ws.row_dimensions[fila].height = 28


def escribir_campo(ws, fila, etiqueta, valor, columnas=2):
    ws.cell(row=fila, column=1, value=etiqueta).font = CAMPOS_FONT
    ws.cell(row=fila, column=1).fill = CAMPOS_FILL
    ws.cell(row=fila, column=1).alignment = WRAP
    ws.cell(row=fila, column=1).border = BORDER_THIN
    celda_valor = ws.cell(row=fila, column=2, value=valor)
    celda_valor.alignment = WRAP
    celda_valor.border = BORDER_THIN
    if columnas == 4:
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=4)
    ws.row_dimensions[fila].height = max(20, min(140, 18 * (1 + len(str(valor)) // 80)))


def ajustar_columnas(ws, ancho_a=8, ancho_b=100):
    ws.column_dimensions[get_column_letter(1)].width = ancho_a
    ws.column_dimensions[get_column_letter(2)].width = ancho_b


# ---- Definicion de los 8 casos de uso ----
AUTORES = "Solano Campos, Luis Francisco; Saldarriaga Mejias, Bruno Favian"
VERSION = "1.0"
OBJ_01 = "OBJ-01: Reduccion de Tiempos de Respuesta al Cliente."
OBJ_02 = "OBJ-02: Estandarizacion Tecnica del 100% de las Fichas Tecnicas."
OBJ_03 = "OBJ-03: Validacion Automatica de Compatibilidad entre Componentes."
RF_01 = "RF-01: Acceso a la Aplicacion."

CASOS = [
    {
        "id": "RF-02",
        "titulo": "RF-02  Gestionar Inventario",
        "objetivos": OBJ_02,
        "requisitos": RF_01,
        "descripcion": "El sistema debera comportarse como se describe en el siguiente caso de uso cuando un administrador decide gestionar el inventario de componentes del catalogo.",
        "precondicion": "El administrador ha iniciado sesion en la aplicacion con credenciales validas y permisos elevados.",
        "secuencia": [
            ("1", "El administrador accede al modulo de gestion de inventario desde el panel principal."),
            ("2", "El sistema muestra la lista de componentes y perifericos registrados con sus datos tecnicos basicos."),
            ("3", "El administrador selecciona una operacion: crear, editar o eliminar un componente."),
            ("4", "El sistema solicita los datos tecnicos del componente (nombre, categoria, socket, TDP, precio base, stock)."),
            ("5", "El administrador ingresa los datos del componente."),
            ("6", "El sistema valida la informacion tecnica ingresada."),
            ("7", "El sistema registra el componente en la base de datos PostgreSQL."),
            ("8", "El sistema confirma la operacion exitosamente y actualiza la lista visible."),
        ],
        "excepciones": [
            ("6", "Si algun dato tecnico es invalido o esta vacio, el sistema muestra un mensaje de error. Ir al paso 4."),
            ("7", "Si ocurre un error en la base de datos, el sistema muestra un mensaje. Ir al paso 2."),
        ],
        "postcondicion": "El inventario queda actualizado y los cambios quedan disponibles para cotizaciones futuras.",
        "importancia": "Vital",
        "urgencia": "Inmediatamente",
        "comentarios": "Ninguno. Carga inicial realizada con 873 componentes y perifericos (Capitulo IV, seccion 4.5.2).",
    },
    {
        "id": "RF-03",
        "titulo": "RF-03  Actualizar Precios",
        "objetivos": OBJ_01 + "; " + OBJ_02,
        "requisitos": RF_01 + "; RF-02: Gestionar Inventario.",
        "descripcion": "El sistema debera permitir al administrador actualizar los precios base, margenes de ganancia e impuestos aplicados a los componentes del catalogo.",
        "precondicion": "El administrador ha iniciado sesion y existe al menos un componente registrado en el inventario.",
        "secuencia": [
            ("1", "El administrador accede al modulo de configuracion de precios."),
            ("2", "El sistema muestra la lista de componentes con sus precios actuales y margenes aplicados."),
            ("3", "El administrador selecciona un componente o aplica un cambio masivo por categoria."),
            ("4", "El sistema muestra el detalle del precio base, margen e impuesto del componente."),
            ("5", "El administrador actualiza el precio base, el margen o el impuesto."),
            ("6", "El sistema valida que los nuevos valores sean numericos y positivos."),
            ("7", "El sistema registra la actualizacion en la base de datos."),
            ("8", "El sistema confirma la operacion y muestra el nuevo precio calculado."),
        ],
        "excepciones": [
            ("6", "Si el valor ingresado no es numerico o es negativo, el sistema muestra un mensaje. Ir al paso 5."),
            ("7", "Si ocurre un error en la base de datos, el sistema muestra un mensaje. Ir al paso 2."),
        ],
        "postcondicion": "Los precios y margenes quedan actualizados y disponibles para las nuevas cotizaciones generadas.",
        "importancia": "Importante",
        "urgencia": "Inmediatamente",
        "comentarios": "Forma parte del protocolo de sostenibilidad operativa: actualizacion semanal de precios y stock (Capitulo IV, seccion 4.2.2).",
    },
    {
        "id": "RF-04",
        "titulo": "RF-04  Crear Cotizacion",
        "objetivos": OBJ_01 + "; " + OBJ_03,
        "requisitos": RF_01 + "; RF-02: Gestionar Inventario; RF-09: Consultar Catalogo.",
        "descripcion": "El sistema debera permitir al vendedor (o administrador) crear una cotizacion tecnica de componentes para un cliente, integrando validacion de compatibilidad y calculo automatizado de precios.",
        "precondicion": "El usuario ha iniciado sesion y existe al menos un componente registrado en el catalogo.",
        "secuencia": [
            ("1", "El usuario accede al modulo de cotizacion desde el panel principal."),
            ("2", "El sistema muestra el catalogo de componentes disponibles filtrables por categoria y socket."),
            ("3", "El usuario selecciona los componentes que conformaran la cotizacion."),
            ("4", "El sistema valida automaticamente la compatibilidad tecnica entre los componentes seleccionados (Socket, RAM, TDP)."),
            ("5", "El sistema calcula el precio total aplicando margenes de ganancia, descuentos e impuestos."),
            ("6", "El usuario revisa el resumen financiero y los datos del cliente."),
            ("7", "El usuario confirma la cotizacion."),
            ("8", "El sistema registra la cotizacion en la base de datos con un codigo unico."),
            ("9", "El sistema muestra la cotizacion generada y la opcion de generar el reporte PDF."),
        ],
        "excepciones": [
            ("4", "Si los componentes seleccionados son incompatibles, el sistema muestra una alerta y bloquea el avance. Ir al paso 3."),
            ("7", "Si ocurre un error en la base de datos, el sistema muestra un mensaje. Ir al paso 2."),
        ],
        "postcondicion": "La cotizacion queda registrada, validada y disponible para generar el reporte PDF o continuar su edicion.",
        "importancia": "Vital",
        "urgencia": "Inmediatamente",
        "comentarios": "Caso central del sistema. Logro reducir el tiempo de emision de proformas de 66 a menos de 2 minutos (Capitulo IV, seccion 4.5.4).",
    },
    {
        "id": "RF-05",
        "titulo": "RF-05  Solicitar Recomendacion de IA",
        "objetivos": OBJ_01 + "; " + OBJ_02 + "; " + OBJ_03,
        "requisitos": RF_01 + "; RF-04: Crear Cotizacion.",
        "descripcion": "El sistema debera permitir al vendedor solicitar al asistente de inteligencia artificial recomendaciones de componentes segun presupuesto, uso previsto y compatibilidad requerida.",
        "precondicion": "El vendedor ha iniciado sesion y el catalogo de componentes esta cargado en la base de datos.",
        "secuencia": [
            ("1", "El vendedor accede al modulo del asistente de IA desde el cotizador."),
            ("2", "El sistema muestra la interfaz conversacional del asistente."),
            ("3", "El vendedor ingresa una consulta en lenguaje natural describiendo presupuesto, uso previsto y preferencias."),
            ("4", "El sistema envia la consulta al Servicio de IA externo (Google Gemini 2.5 Flash)."),
            ("5", "El Servicio de IA procesa la consulta, clasifica los componentes (Llama 3.2) y reordena resultados (Rerank QA Mistral)."),
            ("6", "El sistema recibe y muestra al vendedor las recomendaciones con su justificacion tecnica."),
            ("7", "El vendedor selecciona los componentes sugeridos para incluirlos en la cotizacion."),
            ("8", "El sistema registra la seleccion y redirige al modulo de cotizacion."),
        ],
        "excepciones": [
            ("5", "Si el Servicio de IA principal no responde, el sistema activa el modelo de respaldo (Mistral Small). Volver al paso 4."),
            ("6", "Si la consulta no genera resultados relevantes, el sistema sugiere reformular la consulta. Volver al paso 3."),
        ],
        "postcondicion": "El vendedor recibe recomendaciones tecnicas validadas, aplicables a su cotizacion.",
        "importancia": "Importante",
        "urgencia": "Inmediatamente",
        "comentarios": "Integra la arquitectura multimodelo: Gemini 2.5 Flash (generativo), Llama 3.2 (clasificacion), NV-Embed (embeddings), Rerank QA Mistral (reordenamiento) y Mistral Small (respaldo).",
    },
    {
        "id": "RF-06",
        "titulo": "RF-06  Generar Reporte PDF",
        "objetivos": OBJ_02,
        "requisitos": RF_01 + "; RF-04: Crear Cotizacion.",
        "descripcion": "El sistema debera generar automaticamente un documento PDF profesional con los detalles de la cotizacion, incluyendo identidad corporativa, codigo unico, fecha de caducidad y especificaciones tecnicas.",
        "precondicion": "Existe una cotizacion registrada y confirmada en el sistema.",
        "secuencia": [
            ("1", "El vendedor accede al detalle de la cotizacion confirmada."),
            ("2", "El sistema muestra la opcion de generar el reporte PDF."),
            ("3", "El vendedor solicita la generacion del reporte PDF."),
            ("4", "El sistema compila la informacion de la cotizacion (cliente, componentes, precios, margenes)."),
            ("5", "El sistema genera el documento PDF con la libreria PDFKit en el backend Node.js."),
            ("6", "El sistema incluye: logotipo, datos de contacto, codigo unico, fecha de caducidad y fichas tecnicas."),
            ("7", "El sistema almacena el documento generado y notifica su disponibilidad."),
            ("8", "El sistema ofrece la descarga del PDF al vendedor para su entrega al cliente."),
        ],
        "excepciones": [
            ("5", "Si ocurre un error en la generacion del PDF, el sistema muestra un mensaje. Volver al paso 3."),
            ("8", "Si la descarga falla por conexion, el sistema permite reintentar la operacion."),
        ],
        "postcondicion": "El documento PDF queda disponible para descarga y entrega al cliente.",
        "importancia": "Vital",
        "urgencia": "Inmediatamente",
        "comentarios": "El Cliente recibe y descarga el PDF pero no lo genera. Generacion automatizada mediante PDFKit en el backend (Capitulo IV, seccion 4.1.4).",
    },
    {
        "id": "RF-07",
        "titulo": "RF-07  Visualizar Dashboard",
        "objetivos": OBJ_01,
        "requisitos": RF_01,
        "descripcion": "El sistema debera permitir al administrador y al vendedor visualizar un panel con indicadores de gestion y metricas del proceso de cotizacion.",
        "precondicion": "El usuario ha iniciado sesion en el sistema con credenciales validas.",
        "secuencia": [
            ("1", "El usuario accede al modulo de dashboard desde el panel principal."),
            ("2", "El sistema recupera las metricas operativas del sistema."),
            ("3", "El sistema muestra: total de cotizaciones, cotizaciones pendientes, ventas del periodo y componentes mas cotizados."),
            ("4", "El usuario aplica filtros por periodo de tiempo o categoria de producto."),
            ("5", "El sistema actualiza los indicadores mostrados segun los filtros aplicados."),
        ],
        "excepciones": [
            ("2", "Si ocurre un error al recuperar las metricas, el sistema muestra un mensaje. Ir al paso 1."),
            ("5", "Si los filtros no retornan datos en el periodo, el sistema muestra un estado vacio."),
        ],
        "postcondicion": "El usuario visualiza indicadores actualizados del desempeno del sistema.",
        "importancia": "Importante",
        "urgencia": "Postergado",
        "comentarios": "Caso pendiente de implementacion y documentacion explicita en la tesis. Se incluye en el diagrama de casos de uso como funcionalidad prevista.",
    },
    {
        "id": "RF-08",
        "titulo": "RF-08  Visualizar / Recibir Cotizacion",
        "objetivos": OBJ_01,
        "requisitos": RF_01 + "; RF-04: Crear Cotizacion.",
        "descripcion": "El sistema debera permitir al vendedor visualizar el detalle de las cotizaciones generadas y al cliente recibir y descargar la cotizacion en formato PDF.",
        "precondicion": "Existe al menos una cotizacion registrada en el sistema.",
        "secuencia": [
            ("1", "El usuario accede al historial o modulo de cotizaciones."),
            ("2", "El sistema muestra la lista de cotizaciones disponibles con su estado."),
            ("3", "El usuario selecciona una cotizacion de la lista."),
            ("4", "El sistema muestra el detalle completo de la cotizacion."),
            ("5", "El usuario descarga el PDF asociado (cliente) o continua con el seguimiento comercial (vendedor)."),
        ],
        "excepciones": [
            ("2", "Si no hay cotizaciones registradas, el sistema muestra un estado vacio."),
            ("4", "Si la cotizacion seleccionada ya no existe, el sistema muestra un mensaje. Volver al paso 2."),
        ],
        "postcondicion": "El usuario visualiza o descarga los detalles completos de la cotizacion.",
        "importancia": "Vital",
        "urgencia": "Inmediatamente",
        "comentarios": "Beneficio tecnico no cuantificable documentado: historial de cotizaciones para NSG LATINOAMERICA E.I.R.L. (Capitulo VI, seccion 6.1).",
    },
    {
        "id": "RF-09",
        "titulo": "RF-09  Consultar Catalogo",
        "objetivos": OBJ_02,
        "requisitos": RF_01,
        "descripcion": "El sistema debera permitir al vendedor y al cliente consultar el catalogo de componentes y perifericos disponibles, aplicando filtros por categoria, marca, precio y compatibilidad.",
        "precondicion": "El usuario ha iniciado sesion y el catalogo de componentes esta cargado en la base de datos.",
        "secuencia": [
            ("1", "El usuario accede al modulo de catalogo desde el panel principal."),
            ("2", "El sistema muestra la lista de componentes disponibles con paginacion."),
            ("3", "El usuario aplica filtros por categoria, marca, rango de precio o tipo de socket."),
            ("4", "El sistema muestra los componentes que cumplen los criterios de busqueda."),
            ("5", "El usuario selecciona un componente de la lista."),
            ("6", "El sistema muestra la ficha tecnica detallada del componente."),
        ],
        "excepciones": [
            ("4", "Si ningun componente cumple los filtros, el sistema muestra un estado vacio."),
            ("6", "Si el componente seleccionado no existe, el sistema muestra un mensaje. Volver al paso 2."),
        ],
        "postcondicion": "El usuario obtiene la informacion tecnica y comercial del componente consultado.",
        "importancia": "Vital",
        "urgencia": "Inmediatamente",
        "comentarios": "Catalogo centralizado con 873 componentes y perifericos cargados inicialmente (Capitulo IV, seccion 4.5.2).",
    },
]


def construir_hoja(wb, caso):
    ws = wb.create_sheet(title=caso["id"])
    ajustar_columnas(ws, ancho_a=10, ancho_b=90)
    ws.column_dimensions[get_column_letter(3)].width = 10
    ws.column_dimensions[get_column_letter(4)].width = 90

    fila = 1
    estilizar_titulo(ws, caso["titulo"], fila=fila, columnas=4)
    fila += 1

    escribir_campo(ws, fila, "Version", VERSION, columnas=4); fila += 1
    escribir_campo(ws, fila, "Autores", AUTORES, columnas=4); fila += 1
    escribir_campo(ws, fila, "Objetivos Asociados", caso["objetivos"], columnas=4); fila += 1
    escribir_campo(ws, fila, "Requisitos asociados", caso["requisitos"], columnas=4); fila += 1
    escribir_campo(ws, fila, "Descripcion", caso["descripcion"], columnas=4); fila += 1
    escribir_campo(ws, fila, "Precondicion", caso["precondicion"], columnas=4); fila += 1

    # Encabezado de seccion: Secuencia normal
    ws.cell(row=fila, column=1, value="Secuencia normal").font = CAMPOS_FONT
    ws.cell(row=fila, column=1).fill = CAMPOS_FILL
    ws.cell(row=fila, column=1).alignment = WRAP
    ws.cell(row=fila, column=1).border = BORDER_THIN
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=4)
    ws.cell(row=fila, column=2).border = BORDER_THIN
    ws.row_dimensions[fila].height = 22
    fila += 1

    # Sub-encabezados
    for col, texto in [(1, "Paso"), (2, "Accion")]:
        c = ws.cell(row=fila, column=col, value=texto)
        c.font = HEADER_TABLA_FONT
        c.fill = HEADER_TABLA_FILL
        c.alignment = CENTER
        c.border = BORDER_THIN
    fila += 1

    for paso, accion in caso["secuencia"]:
        c1 = ws.cell(row=fila, column=1, value=paso)
        c1.alignment = CENTER
        c1.border = BORDER_THIN
        c2 = ws.cell(row=fila, column=2, value=accion)
        c2.alignment = WRAP
        c2.border = BORDER_THIN
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=4)
        ws.row_dimensions[fila].height = 24
        fila += 1

    # Excepciones
    ws.cell(row=fila, column=1, value="Excepciones").font = CAMPOS_FONT
    ws.cell(row=fila, column=1).fill = CAMPOS_FILL
    ws.cell(row=fila, column=1).alignment = WRAP
    ws.cell(row=fila, column=1).border = BORDER_THIN
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=4)
    ws.cell(row=fila, column=2).border = BORDER_THIN
    ws.row_dimensions[fila].height = 22
    fila += 1

    for col, texto in [(1, "Paso"), (2, "Accion")]:
        c = ws.cell(row=fila, column=col, value=texto)
        c.font = HEADER_TABLA_FONT
        c.fill = HEADER_TABLA_FILL
        c.alignment = CENTER
        c.border = BORDER_THIN
    fila += 1

    for paso, accion in caso["excepciones"]:
        c1 = ws.cell(row=fila, column=1, value=paso)
        c1.alignment = CENTER
        c1.border = BORDER_THIN
        c2 = ws.cell(row=fila, column=2, value=accion)
        c2.alignment = WRAP
        c2.border = BORDER_THIN
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=4)
        ws.row_dimensions[fila].height = 32
        fila += 1

    escribir_campo(ws, fila, "Postcondicion", caso["postcondicion"], columnas=4); fila += 1
    escribir_campo(ws, fila, "Importancia", caso["importancia"], columnas=4); fila += 1
    escribir_campo(ws, fila, "Urgencia", caso["urgencia"], columnas=4); fila += 1
    escribir_campo(ws, fila, "Comentarios", caso["comentarios"], columnas=4); fila += 1


def main():
    wb = Workbook()
    wb.remove(wb.active)

    for caso in CASOS:
        construir_hoja(wb, caso)

    # Hoja indice al inicio
    indice = wb.create_sheet(title="Indice", index=0)
    indice.column_dimensions[get_column_letter(1)].width = 12
    indice.column_dimensions[get_column_letter(2)].width = 90
    estilizar_titulo(indice, "Documentacion de Casos de Uso - NSG Cotizador", fila=1, columnas=2)
    indice.cell(row=2, column=1, value="Sistema").font = CAMPOS_FONT
    indice.cell(row=2, column=1).fill = CAMPOS_FILL
    indice.cell(row=2, column=1).border = BORDER_THIN
    indice.cell(row=2, column=2, value="NSG Cotizador - Sistema Semiautomatizado de Cotizacion y Catalogacion de Componentes").alignment = WRAP
    indice.cell(row=2, column=2).border = BORDER_THIN
    indice.cell(row=3, column=1, value="Empresa").font = CAMPOS_FONT
    indice.cell(row=3, column=1).fill = CAMPOS_FILL
    indice.cell(row=3, column=1).border = BORDER_THIN
    indice.cell(row=3, column=2, value="NSG LATINOAMERICA E.I.R.L.").alignment = WRAP
    indice.cell(row=3, column=2).border = BORDER_THIN
    indice.cell(row=4, column=1, value="Autores").font = CAMPOS_FONT
    indice.cell(row=4, column=1).fill = CAMPOS_FILL
    indice.cell(row=4, column=1).border = BORDER_THIN
    indice.cell(row=4, column=2, value=AUTORES).alignment = WRAP
    indice.cell(row=4, column=2).border = BORDER_THIN
    indice.cell(row=5, column=1, value="Version").font = CAMPOS_FONT
    indice.cell(row=5, column=1).fill = CAMPOS_FILL
    indice.cell(row=5, column=1).border = BORDER_THIN
    indice.cell(row=5, column=2, value=VERSION).alignment = WRAP
    indice.cell(row=5, column=2).border = BORDER_THIN
    indice.cell(row=6, column=1, value="Plantilla").font = CAMPOS_FONT
    indice.cell(row=6, column=1).fill = CAMPOS_FILL
    indice.cell(row=6, column=1).border = BORDER_THIN
    indice.cell(row=6, column=2, value="RF-01 Acceso Aplicacion").alignment = WRAP
    indice.cell(row=6, column=2).border = BORDER_THIN

    # Encabezados de tabla indice
    fila = 8
    c1 = indice.cell(row=fila, column=1, value="ID")
    c1.font = HEADER_TABLA_FONT
    c1.fill = HEADER_TABLA_FILL
    c1.alignment = CENTER
    c1.border = BORDER_THIN
    c2 = indice.cell(row=fila, column=2, value="Caso de Uso")
    c2.font = HEADER_TABLA_FONT
    c2.fill = HEADER_TABLA_FILL
    c2.alignment = CENTER
    c2.border = BORDER_THIN
    fila += 1
    for caso in CASOS:
        nombre = caso["titulo"].split("  ", 1)[1]
        c1 = indice.cell(row=fila, column=1, value=caso["id"])
        c1.alignment = CENTER
        c1.border = BORDER_THIN
        c2 = indice.cell(row=fila, column=2, value=nombre)
        c2.alignment = WRAP
        c2.border = BORDER_THIN
        # Hipervinculo a la hoja
        c2.hyperlink = "#'" + caso["id"] + "'!A1"
        c2.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
        fila += 1

    ruta_salida = "C:/Users/ASUS/Desktop/Proyecto_Final/NSG_Cotizador/entregables/documentacion-casos-uso/Casos_Uso_NSG_Cotizador.xlsx"
    wb.save(ruta_salida)
    print(f"Archivo generado: {ruta_salida}")
    print(f"Hojas creadas: {len(CASOS) + 1} (1 indice + {len(CASOS)} casos de uso)")


if __name__ == "__main__":
    main()
